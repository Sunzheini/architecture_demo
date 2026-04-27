import openpyxl
wb = openpyxl.load_workbook(r'D:\Study\Projects\Github\architecture_demo\documentation\Tasks.xlsx')
for sheet_name in wb.sheetnames:
    print(f'=== Sheet: {sheet_name} ===')
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print(row)

